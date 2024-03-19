#
#Company Name	Symbol	Qty	Cost/share	Cost
#
import openpyxl
from stock import Stock
from portfolio import Portfolio
import csv

def read_stocks_from_csv(csv_filename: str, start_row: int = 2) -> Portfolio:
    """
    Reads stock data from a CSV file with specific headers and creates a list of Stock objects.

    Args:
        csv_filename (str): The path to the CSV file.
        start_row (int): The row number where the stock data begins (default is 2).

    Returns:
        list[Stock]: A list of Stock objects representing the data in the CSV file.
    """

    stocks = []
    with open(csv_filename, 'r') as file:
        reader = csv.DictReader(file) 

        # Skip to the starting row
        for _ in range(start_row - 1):
            next(reader)

        # Process each data row
        for row in reader:
            try:
                stocks.append(Stock(
                    symbol=row['Symbol'],
                    qty=float(row['Qty']),
                    cost_per_share=float(row['Cost/share']),
                    # No need to calculate 'Cost', as we already have the total cost
                ))
            except ValueError:
                print(f"Error parsing row: {row}")
            except TypeError:
                return Portfolio(stocks)

    return Portfolio(stocks)

def read_stocks_from_excel(excel_filename: str, sheet_name: str = "Sheet1", start_row: int = 2, start_column: str = 'A') -> Portfolio:
    """ 
    Reads stock data from an Excel file.

    Args:
        excel_filename (str): The path to the Excel file.
        sheet_name (str): The name of the sheet containing the data (default is "Sheet1").
        start_row (int): The row number where the stock data begins (default is 2).
        start_column (str): The column letter where the stock data begins (default is 'A').

    Returns:
        list[Stock]: A list of Stock objects representing the data in the Excel file.
    """

    stocks = []
    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook[sheet_name]

    column_index_offset = column_letter_to_index(start_column)
    
    for row_num in range(start_row, sheet.max_row + 1):  # Iterate through rows
        try:
            #print(row_num)
            #print(start_column)
            stocks.append(Stock(
                symbol=sheet.cell(row=row_num, column=column_index_offset + 2).value,
                qty=float(sheet.cell(row=row_num, column=column_index_offset + 3).value),
                cost_per_share=float(sheet.cell(row=row_num, column=column_index_offset + 4).value)
            ))
        except ValueError:
            print(f"Error parsing row: {row_num}")
        except TypeError:
            return Portfolio(stocks)

    return Portfolio(stocks)


def column_letter_to_index(column_letter: str) -> int:
    """Converts a column letter (e.g., 'A', 'B', 'C') to a zero-based index"""
    column_letter = column_letter.upper()
    return ord(column_letter) - ord('A')  # Subtract 65 to get 0 for 'A', 1 for 'B', etc.
